import openpyxl
import json


def get_categories_products(table):
    wb = openpyxl.load_workbook(table)
    ws = wb.active

    hierarchy_data = {}
    row = 2
    enclosure = 0
    first_key = False
    current_key = False
    prod_from_curr_cat = []

    for i in range(ws.max_row):
        unit = ws.cell(row=row, column=3).value

        if not unit:
            category_name = ws.cell(row=row, column=1).value

            row += 1
            enclosure += 1
            if not first_key:
                first_key = category_name
                continue
            else:
                if prod_from_curr_cat:
                    hierarchy_data[current_key] = prod_from_curr_cat
                    prod_from_curr_cat = []

                current_key = category_name
                hierarchy_data[category_name] = None

        else:
            if unit == "кг":
                product = ws.cell(row=row, column=1).value
                prod_from_curr_cat.append(product)
                row += 1
            else:
                row += 1
                continue

    with open(f'{first_key}.json', 'w') as json_file:
        json.dump(hierarchy_data, json_file)
    print(hierarchy_data)


if __name__ == "__main__":
    get_categories_products('Весовой крепеж.xlsx')
